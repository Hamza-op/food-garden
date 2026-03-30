"""
AuraPOS Professional - Database Manager (SQLite)
"""
import sqlite3
from datetime import datetime, timedelta
from typing import Optional, List, Dict, Any, Tuple
import os
from config import DB_PATH, DEFAULT_SETTINGS, BILL_RETENTION_DAYS


class DatabaseManager:
    """Handles all SQLite database operations with robust error handling."""

    MAX_CLOCK_ROLLBACK_PROTECTION_HOURS = 72
    
    def __init__(self, db_path: str = DB_PATH):
        self.db_path = db_path
        self._connection: Optional[sqlite3.Connection] = None
    
    def connect(self) -> sqlite3.Connection:
        """Establish database connection with WAL mode for performance."""
        try:
            # Close existing connection first
            if self._connection:
                try:
                    self._connection.close()
                except Exception:
                    pass
                self._connection = None
            
            self._connection = sqlite3.connect(self.db_path, check_same_thread=False)
            self._connection.row_factory = sqlite3.Row
            self._connection.execute("PRAGMA journal_mode=WAL")
            self._connection.execute("PRAGMA foreign_keys=ON")
            return self._connection
        except sqlite3.Error as e:
            raise RuntimeError(f"Database connection failed: {e}")
    
    def close(self):
        """Close database connection safely."""
        if self._connection:
            try:
                self._connection.commit()  # Commit any pending changes
                self._connection.close()
            except Exception as e:
                print(f"Warning closing connection: {e}")
            self._connection = None
    
    @property
    def connection(self) -> sqlite3.Connection:
        """Get or create database connection."""
        if self._connection is None:
            self.connect()
        return self._connection
    
    def initialize_database(self):
        """Create all required tables if they don't exist."""
        try:
            cursor = self.connection.cursor()
            
            # Users table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS users (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    username TEXT UNIQUE NOT NULL,
                    password_hash TEXT NOT NULL,
                    role TEXT NOT NULL CHECK(role IN ('Admin', 'Staff')),
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)
            
            # Menu table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS menu (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT NOT NULL,
                    category TEXT NOT NULL DEFAULT 'General',
                    price REAL NOT NULL DEFAULT 0,
                    tax_rate REAL DEFAULT 0,
                    status TEXT DEFAULT 'active' CHECK(status IN ('active', 'inactive'))
                )
            """)
            
            # Sales table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS sales (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    receipt_no TEXT UNIQUE NOT NULL,
                    subtotal REAL NOT NULL,
                    tax REAL NOT NULL,
                    discount REAL DEFAULT 0,
                    total REAL NOT NULL,
                    payment_type TEXT NOT NULL,
                    timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    user_id INTEGER,
                    FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE SET NULL
                )
            """)
            
            # Sale Items table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS sale_items (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    sale_id INTEGER NOT NULL,
                    product_id INTEGER NOT NULL,
                    product_name TEXT NOT NULL,
                    qty INTEGER NOT NULL,
                    price_at_sale REAL NOT NULL,
                    FOREIGN KEY (sale_id) REFERENCES sales(id) ON DELETE CASCADE,
                    FOREIGN KEY (product_id) REFERENCES menu(id) ON DELETE CASCADE
                )
            """)

            # Expenses table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS expenses (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    description TEXT NOT NULL,
                    amount REAL NOT NULL,
                    category TEXT DEFAULT 'General',
                    date DATE DEFAULT CURRENT_DATE,
                    timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    user_id INTEGER,
                    FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE SET NULL
                )
            """)
            
            # Settings table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS settings (
                    key TEXT PRIMARY KEY,
                    value TEXT
                )
            """)
            
            # Archived Sales table (for bills older than retention period)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS archived_sales (
                    id INTEGER PRIMARY KEY,
                    receipt_no TEXT NOT NULL,
                    subtotal REAL NOT NULL,
                    tax REAL NOT NULL,
                    discount REAL DEFAULT 0,
                    total REAL NOT NULL,
                    payment_type TEXT NOT NULL,
                    timestamp TIMESTAMP NOT NULL,
                    user_id INTEGER,
                    archived_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)
            
            # Archived Sale Items table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS archived_sale_items (
                    id INTEGER PRIMARY KEY,
                    sale_id INTEGER NOT NULL,
                    product_id INTEGER NOT NULL,
                    product_name TEXT NOT NULL,
                    qty INTEGER NOT NULL,
                    price_at_sale REAL NOT NULL,
                    FOREIGN KEY (sale_id) REFERENCES archived_sales(id) ON DELETE CASCADE
                )
            """)
            
            self.connection.commit()
            self._initialize_default_settings()
            self._create_default_admin()
            
        except sqlite3.Error as e:
            raise RuntimeError(f"Database initialization failed: {e}")
    
    
    def _initialize_default_settings(self):
        """Insert default settings if not present."""
        try:
            cursor = self.connection.cursor()
            for key, value in DEFAULT_SETTINGS.items():
                cursor.execute(
                    "INSERT OR IGNORE INTO settings (key, value) VALUES (?, ?)",
                    (key, str(value))
                )
            self.connection.commit()
        except sqlite3.Error as e:
            print(f"Warning: Failed to initialize settings: {e}")
    
    def _create_default_admin(self):
        """Create default admin user if no users exist."""
        try:
            cursor = self.connection.cursor()
            cursor.execute("SELECT COUNT(*) FROM users")
            if cursor.fetchone()[0] == 0:
                import bcrypt
                default_password = os.environ.get("FOOD_GARDEN_DEFAULT_ADMIN_PASSWORD", "admin")
                password_hash = bcrypt.hashpw(default_password.encode(), bcrypt.gensalt()).decode()
                cursor.execute(
                    "INSERT INTO users (username, password_hash, role) VALUES (?, ?, ?)",
                    ("admin", password_hash, "Admin")
                )
                self.connection.commit()
        except sqlite3.Error as e:
            print(f"Warning: Failed to create default admin: {e}")
    
    # ==================== User Operations ====================
    
    def get_user(self, username: str) -> Optional[Dict[str, Any]]:
        """Get user by username."""
        try:
            cursor = self.connection.cursor()
            cursor.execute("SELECT * FROM users WHERE username = ?", (username,))
            row = cursor.fetchone()
            return dict(row) if row else None
        except sqlite3.Error as e:
            print(f"Error getting user: {e}")
            return None
    
    def add_user(self, username: str, password_hash: str, role: str) -> bool:
        """Add a new user."""
        try:
            cursor = self.connection.cursor()
            cursor.execute(
                "INSERT INTO users (username, password_hash, role) VALUES (?, ?, ?)",
                (username, password_hash, role)
            )
            self.connection.commit()
            return True
        except sqlite3.Error as e:
            print(f"Error adding user: {e}")
            return False
    
    def get_all_users(self) -> List[Dict[str, Any]]:
        """Get all users."""
        try:
            cursor = self.connection.cursor()
            cursor.execute("SELECT id, username, role, created_at FROM users")
            return [dict(row) for row in cursor.fetchall()]
        except sqlite3.Error as e:
            print(f"Error getting users: {e}")
            return []
    
    def delete_user(self, user_id: int) -> bool:
        """Delete a user by ID."""
        try:
            cursor = self.connection.cursor()
            cursor.execute("DELETE FROM users WHERE id = ?", (user_id,))
            self.connection.commit()
            return cursor.rowcount > 0
        except sqlite3.Error as e:
            print(f"Error deleting user: {e}")
            return False
    
    # ==================== Menu Operations ====================
    
    def get_all_menu_items(self, active_only: bool = True) -> List[Dict[str, Any]]:
        """Get all menu items."""
        try:
            cursor = self.connection.cursor()
            if active_only:
                cursor.execute("SELECT * FROM menu WHERE status = 'active' ORDER BY category, name")
            else:
                cursor.execute("SELECT * FROM menu ORDER BY category, name")
            return [dict(row) for row in cursor.fetchall()]
        except sqlite3.Error as e:
            print(f"Error getting menu items: {e}")
            return []
    
    def add_menu_item(self, name: str, category: str, price: float, tax_rate: float = 0) -> bool:
        """Add a new menu item."""
        try:
            if not name or not name.strip():
                print("Error: Item name cannot be empty")
                return False
            
            if not category or not category.strip():
                category = "General"
            
            cursor = self.connection.cursor()
            cursor.execute(
                "INSERT INTO menu (name, category, price, tax_rate, status) VALUES (?, ?, ?, ?, 'active')",
                (name.strip(), category.strip(), float(price), float(tax_rate))
            )
            self.connection.commit()
            print(f"Menu item '{name}' added successfully")
            return True
        except sqlite3.Error as e:
            print(f"Error adding menu item: {e}")
            return False
        except Exception as e:
            print(f"Unexpected error adding menu item: {e}")
            return False
    
    def update_menu_item(self, item_id: int, name: str, category: str, price: float, tax_rate: float, status: str) -> bool:
        """Update an existing menu item."""
        try:
            cursor = self.connection.cursor()
            cursor.execute(
                "UPDATE menu SET name=?, category=?, price=?, tax_rate=?, status=? WHERE id=?",
                (name, category, price, tax_rate, status, item_id)
            )
            self.connection.commit()
            return cursor.rowcount > 0
        except sqlite3.Error as e:
            print(f"Error updating menu item: {e}")
            return False
    
    def check_item_usage(self, item_id: int) -> int:
        """Check how many times a menu item has been used in sales."""
        try:
            cursor = self.connection.cursor()
            cursor.execute("SELECT COUNT(*) FROM sale_items WHERE product_id = ?", (item_id,))
            return cursor.fetchone()[0]
        except sqlite3.Error as e:
            print(f"Error checking item usage: {e}")
            return 0

    def permanently_delete_menu_item(self, item_id: int) -> bool:
        """Permanently delete a menu item and its sales history."""
        try:
            cursor = self.connection.cursor()
            # First delete from sale_items to satisfy foreign key constraint
            cursor.execute("DELETE FROM sale_items WHERE product_id = ?", (item_id,))
            # Then delete from menu
            cursor.execute("DELETE FROM menu WHERE id = ?", (item_id,))
            self.connection.commit()
            print(f"Item {item_id} and its history permanently deleted.")
            return True
        except sqlite3.Error as e:
            print(f"Error permanently deleting menu item: {e}")
            return False

    def delete_menu_item(self, item_id: int) -> bool:
        """Soft delete a menu item (mark inactive)."""
        try:
            cursor = self.connection.cursor()
            cursor.execute("UPDATE menu SET status = 'inactive' WHERE id = ?", (item_id,))
            self.connection.commit()
            return True
        except sqlite3.Error as e:
            print(f"Error deleting menu item: {e}")
            return False
    
    def search_menu(self, query: str) -> List[Dict[str, Any]]:
        """Search menu items by name."""
        try:
            cursor = self.connection.cursor()
            cursor.execute(
                "SELECT * FROM menu WHERE status = 'active' AND name LIKE ? ORDER BY name",
                (f"%{query}%",)
            )
            return [dict(row) for row in cursor.fetchall()]
        except sqlite3.Error as e:
            print(f"Error searching menu: {e}")
            return []
    
    # ==================== Sales Operations ====================

    def _parse_db_timestamp(self, value: Any) -> Optional[datetime]:
        """Parse app timestamps stored in SQLite."""
        if not value:
            return None

        text = str(value).strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M:%S.%f"):
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                continue
        return None

    def _get_latest_sale_timestamp(self) -> Optional[datetime]:
        """Get the latest recorded sale timestamp from active sales."""
        try:
            cursor = self.connection.cursor()
            cursor.execute(
                "SELECT timestamp FROM sales ORDER BY timestamp DESC, id DESC LIMIT 1"
            )
            row = cursor.fetchone()
            return self._parse_db_timestamp(row[0]) if row else None
        except sqlite3.Error:
            return None

    def _get_safe_sale_datetime(self) -> datetime:
        """
        Return a monotonic local datetime for sales.

        This protects against small backwards clock jumps from a weak laptop CMOS
        battery without permanently anchoring the app to a wildly wrong future date.
        """
        now = datetime.now().replace(microsecond=0)
        latest_sale_time = self._get_latest_sale_timestamp()
        if latest_sale_time is None:
            return now

        if now >= latest_sale_time:
            return now

        rollback_window = latest_sale_time - now
        if rollback_window <= timedelta(hours=self.MAX_CLOCK_ROLLBACK_PROTECTION_HOURS):
            return latest_sale_time + timedelta(seconds=1)

        return now
    
    def generate_receipt_no(self, sale_time: Optional[datetime] = None) -> str:
        """Generate a unique receipt number for the given sale datetime."""
        try:
            if sale_time is None:
                sale_time = self._get_safe_sale_datetime()

            cursor = self.connection.cursor()
            date_str = sale_time.strftime("%Y%m%d")
            cursor.execute(
                """
                SELECT COUNT(*)
                FROM sales
                WHERE receipt_no LIKE ?
                """,
                (f"RCP-{date_str}-%",)
            )
            daily_count = cursor.fetchone()[0]
            return f"RCP-{date_str}-{daily_count + 1:04d}"
        except sqlite3.Error:
            fallback_time = sale_time or datetime.now()
            return f"RCP-{fallback_time.strftime('%Y%m%d%H%M%S')}"
    
    def create_sale(self, subtotal: float, tax: float, discount: float, total: float,
                    payment_type: str, user_id: int, items: List[Dict]) -> Optional[int]:
        """Create a new sale with items."""
        try:
            cursor = self.connection.cursor()
            sale_time = self._get_safe_sale_datetime()
            receipt_no = self.generate_receipt_no(sale_time)
            ts_local = sale_time.strftime("%Y-%m-%d %H:%M:%S")
            
            cursor.execute(
                """INSERT INTO sales (receipt_no, subtotal, tax, discount, total, payment_type, timestamp, user_id)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                (receipt_no, subtotal, tax, discount, total, payment_type, ts_local, user_id)
            )
            sale_id = cursor.lastrowid
            
            for item in items:
                cursor.execute(
                    """INSERT INTO sale_items (sale_id, product_id, product_name, qty, price_at_sale)
                       VALUES (?, ?, ?, ?, ?)""",
                    (sale_id, item["product_id"], item["product_name"], item["qty"], item["price"])
                )
            
            self.connection.commit()
            return sale_id
        except sqlite3.Error as e:
            print(f"Error creating sale: {e}")
            try:
                self.connection.rollback()
            except Exception:
                pass
            return None
    
    def get_sale(self, sale_id: int) -> Optional[Dict[str, Any]]:
        """Get sale by ID with items and cashier name."""
        try:
            cursor = self.connection.cursor()
            # Fetch sale with username
            query = """
                SELECT s.*, u.username as cashier_name 
                FROM sales s 
                LEFT JOIN users u ON s.user_id = u.id 
                WHERE s.id = ?
            """
            cursor.execute(query, (sale_id,))
            sale = cursor.fetchone()
            if not sale:
                return None
            
            sale_dict = dict(sale)
            # Map cashier_name to cashier key for consistency
            sale_dict['cashier'] = sale_dict.get('cashier_name', 'Unknown')
            
            cursor.execute("SELECT * FROM sale_items WHERE sale_id = ?", (sale_id,))
            sale_dict["items"] = [dict(row) for row in cursor.fetchall()]
            return sale_dict
        except sqlite3.Error as e:
            print(f"Error getting sale: {e}")
            return None
    
    def get_last_sale(self) -> Optional[Dict[str, Any]]:
        """Get the most recent sale."""
        try:
            cursor = self.connection.cursor()
            cursor.execute("SELECT id FROM sales ORDER BY id DESC LIMIT 1")
            row = cursor.fetchone()
            if row:
                return self.get_sale(row[0])
            return None
        except sqlite3.Error as e:
            print(f"Error getting last sale: {e}")
            return None
    
    def get_daily_summary(self, date: Optional[str] = None) -> Dict[str, Any]:
        """Get daily sales summary."""
        try:
            if date is None:
                date = datetime.now().strftime("%Y-%m-%d")
            
            cursor = self.connection.cursor()
            
            # Total sales
            cursor.execute(
                """SELECT COUNT(*) as count, COALESCE(SUM(total), 0) as total,
                          COALESCE(SUM(tax), 0) as tax, COALESCE(SUM(discount), 0) as discount
                   FROM sales WHERE DATE(timestamp) = ?""",
                (date,)
            )
            summary = dict(cursor.fetchone())
            
            # By payment type
            cursor.execute(
                """SELECT payment_type, COUNT(*) as count, SUM(total) as total
                   FROM sales WHERE DATE(timestamp) = ? GROUP BY payment_type""",
                (date,)
            )
            summary["by_payment"] = [dict(row) for row in cursor.fetchall()]
            
            # Top items
            cursor.execute(
                """SELECT si.product_name, SUM(si.qty) as total_qty
                   FROM sale_items si
                   JOIN sales s ON si.sale_id = s.id
                   WHERE DATE(s.timestamp) = ?
                   GROUP BY si.product_name
                   ORDER BY total_qty DESC LIMIT 10""",
                (date,)
            )
            summary["top_items"] = [dict(row) for row in cursor.fetchall()]
            
            return summary
        except sqlite3.Error as e:
            print(f"Error getting daily summary: {e}")
            return {"count": 0, "total": 0, "tax": 0, "discount": 0, "by_payment": [], "top_items": []}
    
    # ==================== Settings Operations ====================
    
    def get_setting(self, key: str, default=None) -> Optional[str]:
        """Get a setting value."""
        try:
            cursor = self.connection.cursor()
            cursor.execute("SELECT value FROM settings WHERE key = ?", (key,))
            row = cursor.fetchone()
            return row[0] if row else default
        except sqlite3.Error:
            return default
    
    def set_setting(self, key: str, value: str) -> bool:
        """Set a setting value."""
        try:
            cursor = self.connection.cursor()
            cursor.execute(
                "INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)",
                (key, value)
            )
            self.connection.commit()
            return True
        except sqlite3.Error as e:
            print(f"Error setting value: {e}")
            return False
    
    def get_all_settings(self) -> Dict[str, str]:
        """Get all settings."""
        try:
            cursor = self.connection.cursor()
            cursor.execute("SELECT key, value FROM settings")
            return {row[0]: row[1] for row in cursor.fetchall()}
        except sqlite3.Error:
            return {}

    # ==================== Expense Operations ====================

    def add_expense(self, description: str, amount: float, category: str = "General", user_id: int = None, date: Optional[str] = None) -> bool:
        """Add a new expense."""
        try:
            if not category or not category.strip():
                category = "General"
            
            if date is None:
                date = datetime.now().strftime("%Y-%m-%d")

            cursor = self.connection.cursor()
            cursor.execute(
                "INSERT INTO expenses (description, amount, category, user_id, date) VALUES (?, ?, ?, ?, ?)",
                (description, amount, category, user_id, date)
            )
            self.connection.commit()
            return True
        except sqlite3.Error as e:
            print(f"Error adding expense: {e}")
            print(f"Values: desc={description}, amt={amount}, cat={category}, user={user_id}, date={date}")
            raise # Re-raise to let the UI catch it and show detail

    def get_expenses(self, date: Optional[str] = None) -> List[Dict[str, Any]]:
        """Get expenses for a specific date (defaults to today) with local timestamps."""
        try:
            if date is None:
                date = datetime.now().strftime("%Y-%m-%d")
            
            cursor = self.connection.cursor()
            # Convert UTC timestamp to local time for display
            cursor.execute(
                """SELECT *, datetime(timestamp, 'localtime') as local_timestamp 
                   FROM expenses WHERE date = ? ORDER BY timestamp DESC""",
                (date,)
            )
            return [dict(row) for row in cursor.fetchall()]
        except sqlite3.Error as e:
            print(f"Error getting expenses: {e}")
            return []

    def delete_expense(self, expense_id: int) -> bool:
        """Delete an expense."""
        try:
            cursor = self.connection.cursor()
            cursor.execute("DELETE FROM expenses WHERE id = ?", (expense_id,))
            self.connection.commit()
            return True
        except sqlite3.Error as e:
            print(f"Error deleting expense: {e}")
            return False

    def get_daily_profit(self, date: Optional[str] = None) -> Dict[str, float]:
        """Calculate daily profit (Net Sales - Expenses). Net Sales excludes Tax."""
        try:
            if date is None:
                date = datetime.now().strftime("%Y-%m-%d")
            
            sales_summary = self.get_daily_summary(date)
            total_revenue = sales_summary.get("total", 0)  # Inclusive of tax
            total_tax = sales_summary.get("tax", 0)
            net_sales = total_revenue - total_tax
            
            cursor = self.connection.cursor()
            cursor.execute(
                "SELECT COALESCE(SUM(amount), 0) FROM expenses WHERE date = ?",
                (date,)
            )
            total_expenses = cursor.fetchone()[0]
            
            return {
                "total_revenue": total_revenue,
                "total_tax": total_tax,
                "net_sales": net_sales,
                "total_expenses": total_expenses,
                "net_profit": net_sales - total_expenses
            }
        except sqlite3.Error as e:
            print(f"Error calculating profit: {e}")
            return {
                "total_revenue": 0,
                "total_tax": 0,
                "net_sales": 0,
                "total_expenses": 0,
                "net_profit": 0
            }

    # ==================== Bill Retrieval Operations ====================

    def get_bills_by_date_range(self, start_date: str, end_date: str) -> List[Dict[str, Any]]:
        """Get all bills within a date range."""
        try:
            cursor = self.connection.cursor()
            cursor.execute("""
                SELECT s.*, u.username as cashier_name 
                FROM sales s 
                LEFT JOIN users u ON s.user_id = u.id 
                WHERE DATE(s.timestamp) BETWEEN ? AND ?
                ORDER BY s.timestamp DESC
            """, (start_date, end_date))
            bills = []
            for row in cursor.fetchall():
                bill = dict(row)
                bill['cashier'] = bill.get('cashier_name', 'Unknown')
                # Get items for this bill
                cursor.execute("SELECT * FROM sale_items WHERE sale_id = ?", (bill['id'],))
                bill['items'] = [dict(item) for item in cursor.fetchall()]
                bills.append(bill)
            return bills
        except sqlite3.Error as e:
            print(f"Error getting bills by date range: {e}")
            return []

    # ==================== Archive Operations ====================

    def archive_old_bills(self, cutoff_days: int = None) -> Tuple[bool, str, int]:
        """
        Move bills older than cutoff_days to archive tables.
        Returns: (success, message, count_archived)
        """
        if cutoff_days is None:
            cutoff_days = BILL_RETENTION_DAYS
        
        try:
            cursor = self.connection.cursor()
            cutoff_date = (datetime.now() - timedelta(days=cutoff_days)).strftime("%Y-%m-%d")
            
            # Find bills to archive
            cursor.execute("""
                SELECT id FROM sales 
                WHERE DATE(timestamp) < ?
            """, (cutoff_date,))
            old_sale_ids = [row[0] for row in cursor.fetchall()]
            
            if not old_sale_ids:
                return True, "No bills to archive.", 0
            
            archived_count = 0
            for sale_id in old_sale_ids:
                # Move sale to archive
                cursor.execute("""
                    INSERT INTO archived_sales (id, receipt_no, subtotal, tax, discount, total, payment_type, timestamp, user_id)
                    SELECT id, receipt_no, subtotal, tax, discount, total, payment_type, timestamp, user_id
                    FROM sales WHERE id = ?
                """, (sale_id,))
                
                # Move sale items to archive
                cursor.execute("""
                    INSERT INTO archived_sale_items (id, sale_id, product_id, product_name, qty, price_at_sale)
                    SELECT id, sale_id, product_id, product_name, qty, price_at_sale
                    FROM sale_items WHERE sale_id = ?
                """, (sale_id,))
                
                # Delete from original tables (cascade will handle items)
                cursor.execute("DELETE FROM sales WHERE id = ?", (sale_id,))
                archived_count += 1
            
            self.connection.commit()
            return True, f"Successfully archived {archived_count} bills.", archived_count
            
        except sqlite3.Error as e:
            try:
                self.connection.rollback()
            except Exception:
                pass
            return False, f"Archive failed: {str(e)}", 0

    def get_archived_bills(self, start_date: Optional[str] = None, end_date: Optional[str] = None) -> List[Dict[str, Any]]:
        """Get archived bills, optionally filtered by date range."""
        try:
            cursor = self.connection.cursor()
            
            if start_date and end_date:
                cursor.execute("""
                    SELECT * FROM archived_sales 
                    WHERE DATE(timestamp) BETWEEN ? AND ?
                    ORDER BY timestamp DESC
                """, (start_date, end_date))
            else:
                cursor.execute("SELECT * FROM archived_sales ORDER BY timestamp DESC")
            
            bills = []
            for row in cursor.fetchall():
                bill = dict(row)
                cursor.execute("SELECT * FROM archived_sale_items WHERE sale_id = ?", (bill['id'],))
                bill['items'] = [dict(item) for item in cursor.fetchall()]
                bills.append(bill)
            return bills
        except sqlite3.Error as e:
            print(f"Error getting archived bills: {e}")
            return []

    def restore_archived_bill(self, archived_sale_id: int) -> Tuple[bool, str]:
        """Restore a bill from archive to active sales."""
        try:
            cursor = self.connection.cursor()
            
            # Check if bill exists in archive
            cursor.execute("SELECT * FROM archived_sales WHERE id = ?", (archived_sale_id,))
            archived_sale = cursor.fetchone()
            if not archived_sale:
                return False, "Archived bill not found."
            
            # Move back to sales
            cursor.execute("""
                INSERT INTO sales (id, receipt_no, subtotal, tax, discount, total, payment_type, timestamp, user_id)
                SELECT id, receipt_no, subtotal, tax, discount, total, payment_type, timestamp, user_id
                FROM archived_sales WHERE id = ?
            """, (archived_sale_id,))
            
            # Move items back
            cursor.execute("""
                INSERT INTO sale_items (id, sale_id, product_id, product_name, qty, price_at_sale)
                SELECT id, sale_id, product_id, product_name, qty, price_at_sale
                FROM archived_sale_items WHERE sale_id = ?
            """, (archived_sale_id,))
            
            # Delete from archive
            cursor.execute("DELETE FROM archived_sales WHERE id = ?", (archived_sale_id,))
            
            self.connection.commit()
            return True, "Bill restored successfully."
            
        except sqlite3.Error as e:
            try:
                self.connection.rollback()
            except Exception:
                pass
            return False, f"Restore failed: {str(e)}"

    def get_archive_stats(self) -> Dict[str, Any]:
        """Get statistics about archived bills."""
        try:
            cursor = self.connection.cursor()
            cursor.execute("SELECT COUNT(*), COALESCE(SUM(total), 0) FROM archived_sales")
            row = cursor.fetchone()
            cursor.execute("SELECT MIN(timestamp), MAX(timestamp) FROM archived_sales")
            date_row = cursor.fetchone()
            return {
                "count": row[0],
                "total_value": row[1],
                "oldest_date": date_row[0],
                "newest_date": date_row[1]
            }
        except sqlite3.Error as e:
            print(f"Error getting archive stats: {e}")
            return {"count": 0, "total_value": 0, "oldest_date": None, "newest_date": None}

    # ==================== Excel Import Operations ====================

    def import_menu_from_excel(self, file_path: str) -> Tuple[bool, str, List[Dict]]:
        """
        Import menu items from an Excel file.
        Returns: (success, message, list of import results per row)
        
        Supported columns: name/item_name, category, price, tax_rate, status
        Duplicate names will update existing items.
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            return False, "openpyxl library not installed. Run: pip install openpyxl", []
        
        results = []
        
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            sheet = wb.active
            
            if sheet is None:
                return False, "No active sheet found in Excel file.", []
            
            # Get headers from first row
            headers = []
            for cell in sheet[1]:
                headers.append(str(cell.value).lower().strip() if cell.value else "")
            
            # Map column indices
            col_map = {}
            for idx, header in enumerate(headers):
                if header in ('name', 'item_name', 'item name', 'product', 'product_name'):
                    col_map['name'] = idx
                elif header in ('category', 'cat', 'type'):
                    col_map['category'] = idx
                elif header in ('price', 'cost', 'amount'):
                    col_map['price'] = idx
                elif header in ('tax_rate', 'tax', 'tax rate', 'vat'):
                    col_map['tax_rate'] = idx
                elif header in ('status', 'active', 'state'):
                    col_map['status'] = idx
            
            if 'name' not in col_map:
                return False, "Required column 'name' (or 'item_name') not found in Excel file.", []
            if 'price' not in col_map:
                return False, "Required column 'price' not found in Excel file.", []
            
            # Get default tax rate from settings
            default_tax = float(self.get_setting('tax_rate', 0) or 0)
            
            imported = 0
            updated = 0
            skipped = 0
            
            cursor = self.connection.cursor()
            
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                row_result = {"row": row_num, "status": "success", "message": ""}
                
                try:
                    # Extract values
                    name = str(row[col_map['name']]).strip() if row[col_map['name']] else ""
                    
                    if not name:
                        row_result["status"] = "skipped"
                        row_result["message"] = "Empty name"
                        skipped += 1
                        results.append(row_result)
                        continue
                    
                    # Price validation
                    price_raw = row[col_map['price']]
                    try:
                        price = float(price_raw) if price_raw is not None else 0
                    except (ValueError, TypeError):
                        row_result["status"] = "error"
                        row_result["message"] = f"Invalid price: {price_raw}"
                        skipped += 1
                        results.append(row_result)
                        continue
                    
                    if price <= 0:
                        row_result["status"] = "error"
                        row_result["message"] = f"Price must be positive: {price}"
                        skipped += 1
                        results.append(row_result)
                        continue
                    
                    # Optional fields
                    category = "General"
                    cat_idx = col_map.get("category")
                    if cat_idx is not None and 0 <= int(cat_idx) < len(row):
                        cat_raw = row[int(cat_idx)]
                        if cat_raw is not None and str(cat_raw).strip():
                            category = str(cat_raw).strip()
                    
                    tax_rate = default_tax
                    if col_map.get('tax_rate') is not None and row[col_map['tax_rate']] is not None:
                        try:
                            tax_rate = float(row[col_map['tax_rate']])
                        except (ValueError, TypeError):
                            pass  # Use default
                    
                    status = "active"
                    if col_map.get('status') is not None and row[col_map['status']]:
                        status_raw = str(row[col_map['status']]).lower().strip()
                        if status_raw in ('inactive', 'disabled', 'no', 'false', '0'):
                            status = "inactive"
                    
                    # Check if item exists
                    cursor.execute("SELECT id FROM menu WHERE name = ?", (name,))
                    existing = cursor.fetchone()
                    
                    if existing:
                        # Update existing item
                        cursor.execute("""
                            UPDATE menu SET category=?, price=?, tax_rate=?, status=? WHERE id=?
                        """, (category, price, tax_rate, status, existing[0]))
                        row_result["message"] = f"Updated: {name}"
                        updated += 1
                    else:
                        # Insert new item
                        cursor.execute("""
                            INSERT INTO menu (name, category, price, tax_rate, status) 
                            VALUES (?, ?, ?, ?, ?)
                        """, (name, category, price, tax_rate, status))
                        row_result["message"] = f"Imported: {name}"
                        imported += 1
                    
                    results.append(row_result)
                    
                except Exception as e:
                    row_result["status"] = "error"
                    row_result["message"] = str(e)
                    skipped += 1
                    results.append(row_result)
            
            self.connection.commit()
            wb.close()
            
            summary = f"Import complete: {imported} new, {updated} updated, {skipped} skipped."
            return True, summary, results
            
        except Exception as e:
            return False, f"Excel import failed: {str(e)}", results


# Global database instance
db = DatabaseManager()
