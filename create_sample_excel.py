"""
Sample Excel Generator for Menu Import Testing
Run this script to create test Excel files for the menu import feature.
"""
from openpyxl import Workbook
import os

# Get the script's directory
script_dir = os.path.dirname(os.path.abspath(__file__))

def create_sample_menu_excel():
    """Create a sample menu Excel file for testing imports."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Menu Items"
    
    # Headers
    headers = ["name", "category", "price", "tax_rate", "status"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Sample data
    items = [
        ("Chicken Burger", "Fast Food", 350, 5, "active"),
        ("Cheese Pizza", "Fast Food", 450, 5, "active"),
        ("Vegetable Biryani", "Main Course", 280, 5, "active"),
        ("Chicken Tikka", "Starters", 320, 5, "active"),
        ("Mango Lassi", "Beverages", 120, 0, "active"),
        ("Chocolate Brownie", "Desserts", 180, 5, "active"),
        ("French Fries", "Sides", 150, 5, "active"),
        ("Green Salad", "Sides", 100, 0, "active"),
        ("Mineral Water", "Beverages", 50, 0, "active"),
        ("Ice Cream Sundae", "Desserts", 220, 5, "active"),
    ]
    
    for row, item in enumerate(items, 2):
        for col, value in enumerate(item, 1):
            ws.cell(row=row, column=col, value=value)
    
    # Save the file
    filepath = os.path.join(script_dir, "sample_menu.xlsx")
    wb.save(filepath)
    print(f"Created: {filepath}")
    return filepath

def create_minimal_excel():
    """Create an Excel file with only required columns."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Menu"
    
    # Only required headers
    headers = ["item_name", "price"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    items = [
        ("Coffee", 80),
        ("Tea", 50),
        ("Sandwich", 120),
    ]
    
    for row, item in enumerate(items, 2):
        for col, value in enumerate(item, 1):
            ws.cell(row=row, column=col, value=value)
    
    filepath = os.path.join(script_dir, "minimal_menu.xlsx")
    wb.save(filepath)
    print(f"Created: {filepath}")
    return filepath

if __name__ == "__main__":
    create_sample_menu_excel()
    create_minimal_excel()
    print("\nSample Excel files created successfully!")
    print("Use these files to test the 'Import from Excel' feature in the Admin Panel.")
